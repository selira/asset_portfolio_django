import openpyxl
from django.core.management.base import BaseCommand
from apps.api.models import Portfolio, Asset, PortfolioAssetWeight, AssetPrice

class Command(BaseCommand):
    help = 'Reads data from an Excel file with two sheets'

    def handle(self, *args, **options):
        file_path = 'data/datos.xlsx'
        
        try:
            # Load the workbook
            workbook = openpyxl.load_workbook(file_path)
            
            # Get sheet names
            sheet_names = workbook.sheetnames
            if len(sheet_names) != 2:
                self.stdout.write(self.style.ERROR('Excel file must have 2 sheets'))
                return
            
            # Process first sheet
            sheet1 = workbook[sheet_names[0]]
            self.stdout.write(f"Processing sheet: {sheet_names[0]}")
            
            # Read headers from first row
            headers1 = [cell.value for cell in sheet1[1]]
            
            p1 = Portfolio.objects.create(
                name=headers1[2]
            )
            p2 = Portfolio.objects.create(
                name=headers1[3]
            )

            assets = []

            
            for row in sheet1.iter_rows(min_row=2):
                row_data = [cell.value for cell in row]
                if row_data[1] is None:
                    break
                asset = Asset.objects.create(
                    name=row_data[1]
                )
                assets.append(asset)
                PortfolioAssetWeight.objects.create(
                    asset=asset,
                    portfolio=p1,
                    weight=row_data[2],
                    date=row_data[0]
                )
                PortfolioAssetWeight.objects.create(
                    asset=asset,
                    portfolio=p2,
                    weight=row_data[3],
                    date=row_data[0]
                )           

            # Process second sheet
            sheet2 = workbook[sheet_names[1]]
            self.stdout.write(f"\nProcessing sheet: {sheet_names[1]}")
            
            # Read headers from first row
            headers2 = [cell.value for cell in sheet2[1]]
            ordered_assets = []
            for h in headers2:
                for a in assets:
                    if h == a.name:
                        ordered_assets.append(a)

            batch_size = 1000
            price_objects = []
            
            # Read data rows
            for row in sheet2.iter_rows(min_row=2):
                row_data = [cell.value for cell in row]
                for i, asset in enumerate(ordered_assets):
                    price_objects.append(
                    AssetPrice(
                        asset=asset,
                        price=row_data[i + 1],
                        date=row_data[0]
                       )
                     )
                if len(price_objects) >= batch_size:
                    AssetPrice.objects.bulk_create(price_objects)
                    price_objects = []

            if price_objects:
                AssetPrice.objects.bulk_create(price_objects)
            self.stdout.write(self.style.SUCCESS('Data imported successfully'))

        except FileNotFoundError:
            self.stdout.write(self.style.ERROR(f'File not found: {file_path}'))
        except Exception as e:
            self.stdout.write(self.style.ERROR(f'Error processing file: {str(e)}'))